import argparse
import io
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import sdmx
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass(frozen=True)
class IMFSeries:
    name: str
    dataflow: str
    key_suffix: str
    source_url: str


@dataclass(frozen=True)
class BLSSeries:
    name: str
    series_id: str
    source_url: str


DEFAULT_SERIES = (
    IMFSeries(
        name="CPI",
        dataflow="CPI",
        key_suffix="CPI._T.SRP_IX.M",
        source_url="https://data.imf.org/en/datasets/IMF.STA:CPI",
    ),
    IMFSeries(
        name="PPI",
        dataflow="PPI",
        key_suffix="PPI.IX.M",
        source_url="https://data.imf.org/en/datasets/IMF.STA:PPI",
    ),
)

BLS_PPI_SERIES = BLSSeries(
    name="BLS PPI",
    series_id="WPUFD4",
    source_url="https://www.bls.gov/ppi/",
)

BLS_API_KEY = "fa7905c8cd9f49a4a171df26d7f42a37"

OUTPUT_FILENAME_TEMPLATE = "IMF_CPI_PPI_Raw_{date}.xlsx"
COUNTRY_MAPPING_URL = "https://raw.githubusercontent.com/lukes/ISO-3166-Countries-with-Regional-Codes/master/all/all.csv"


def build_key(key_suffix, countries=None):
    country_key = "+".join(countries) if countries else ""
    return f"{country_key}.{key_suffix}"


def get_column_letter(col_num):
    letter = ""
    while col_num > 0:
        col_num -= 1
        letter = chr(65 + (col_num % 26)) + letter
        col_num //= 26
    return letter


def parse_period(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-M%m", "%Y"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def period_sort_key(value):
    parsed = parse_period(value)
    return parsed or datetime.min


def format_period_label(value):
    parsed = parse_period(value)
    if parsed is None:
        return value
    if isinstance(value, str) and "-M" in value:
        return parsed.strftime("%B %Y")
    return parsed.strftime("%Y")


def latest_period(df):
    best = None
    for value in df.index:
        parsed = parse_period(value)
        if parsed is not None and (best is None or parsed > best):
            best = parsed
    return best


def add_most_recent_row(df, label="Most Recent Data"):
    summary = {}

    for col in df.columns:
        col_series = df[col].dropna()
        best_index = None
        best_date = None
        for idx in col_series.index:
            parsed = parse_period(idx)
            if parsed is not None and (best_date is None or parsed > best_date):
                best_date = parsed
                best_index = idx
        summary[col] = format_period_label(best_index) if best_index is not None else None

    return pd.DataFrame([summary], index=[label])


def load_country_mapping():
    response = requests.get(COUNTRY_MAPPING_URL)
    response.raise_for_status()
    df = pd.read_csv(io.StringIO(response.text))
    return df.set_index("alpha-3")["name"].to_dict()


def country_column_name(country_code, country_mapping=None):
    country_code = str(country_code)
    country_name = (country_mapping or {}).get(country_code)
    if country_name:
        return country_name
    return country_code


def fetch_series_frame(client, series, countries=None, start_period=None, end_period=None):
    params = {}
    if start_period:
        params["startPeriod"] = start_period
    if end_period:
        params["endPeriod"] = end_period

    data = client.data(series.dataflow, key=build_key(series.key_suffix, countries), params=params)
    df = sdmx.to_pandas(data).reset_index()
    if df.empty:
        raise RuntimeError(f"No IMF observations returned for {series.name}.")
    return df


def fetch_bls_series(series, start_period=None, end_period=None):
    url = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
    payload = {
        "seriesid": [series.series_id],
        "registrationkey": BLS_API_KEY,
    }

    if start_period:
        payload["startyear"] = start_period[:4]
    if end_period:
        payload["endyear"] = end_period[:4]

    headers = {"Content-type": "application/json"}
    response = requests.post(url, data=json.dumps(payload), headers=headers)
    response.raise_for_status()

    json_data = response.json()
    if json_data["status"] != "REQUEST_SUCCEEDED":
        raise RuntimeError(f"BLS API error: {json_data.get('message')}")

    results = []
    for s in json_data["Results"]["series"]:
        for item in s["data"]:
            results.append(
                {
                    "TIME_PERIOD": f"{item['year']}-{item['period']}",
                    "COUNTRY": "USA",
                    "value": float(item["value"]),
                }
            )

    if not results:
        raise RuntimeError(f"No BLS observations returned for {series.name}.")

    return pd.DataFrame(results)


def pivot_raw_data(df, country_mapping=None):
    required = {"TIME_PERIOD", "COUNTRY", "value"}
    missing = required.difference(df.columns)
    if missing:
        raise RuntimeError(f"IMF response is missing expected columns: {', '.join(sorted(missing))}")

    raw = df.pivot_table(
        index="TIME_PERIOD",
        columns="COUNTRY",
        values="value",
        aggfunc="first",
    )
    ordered_index = sorted(raw.index, key=period_sort_key, reverse=True)
    return raw.loc[ordered_index]


def format_dataset_date(date_value):
    if date_value is None:
        return "unknown"
    return date_value.strftime("%B %Y")


def default_output_path(tables):
    date_stamp = datetime.now().strftime("%Y%m%d")
    return Path(OUTPUT_FILENAME_TEMPLATE.format(date=date_stamp))


def add_country_code_row(df, label="Country Code"):
    codes = {col: col for col in df.columns}
    return pd.DataFrame([codes], index=[label])


def write_raw_excel(tables, output_path=None, generated_at=None, country_mapping=None):
    generated_at = generated_at or datetime.now()
    output_path = Path(output_path) if output_path else default_output_path(tables)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_payloads = []
        for series, raw_df in tables.items():
            summary_rows = pd.concat([add_country_code_row(raw_df), add_most_recent_row(raw_df)])
            export_df = pd.concat([summary_rows, raw_df.copy()])
            export_df = export_df.reset_index().rename(columns={"index": ""})

            def get_year(time_val):
                if time_val in ["Country Code", "Most Recent Data"]:
                    return ""
                p = parse_period(time_val)
                return p.year if p else ""

            export_df.insert(0, "Year", export_df[""].map(get_year))

            sheet_name = f"{series.name} Raw Data"
            export_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            sheet_payloads.append((writer.sheets[sheet_name], series, export_df, latest_period(raw_df)))

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        bold_font = Font(bold=True)
        italic_font = Font(italic=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for ws, series, export_df, latest_date in sheet_payloads:
            source_name = "BLS" if "bls.gov" in series.source_url else "IMF"
            latest_label = format_dataset_date(latest_date)
            ws["A1"] = (
                f"Raw {series.name} index data from {source_name} ({series.source_url}); "
                f"latest observation period in this sheet: {latest_label}; "
                f"file generated {generated_at.strftime('%B %d, %Y')}."
            )
            ws["A1"].font = Font(italic=True, size=9)
            ws["A2"] = f"Values are unrebased {source_name} observations; no model calculations are applied."
            ws["A2"].font = Font(italic=True, size=9)

            # Set Row 3 Headers with Country Names
            for col_num, col_name in enumerate(export_df.columns, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                
                if col_num in [1, 2]:
                    cell.value = ""
                else:
                    # Rename the header cell to full country name
                    cell.value = country_column_name(col_name, country_mapping)

            for row_num, row in enumerate(export_df.values, 4):
                label = row[1]
                is_summary_row = label in ["Country Code", "Most Recent Data"]
                is_country_code_row = label == "Country Code"
                is_most_recent_row = label == "Most Recent Data"
                
                for col_num, value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border
                    cell.alignment = center_alignment

                    if col_num <= 2:  # Year or Time
                        cell.font = Font(bold=True, italic=is_most_recent_row)
                        if col_num == 2 and not is_summary_row:
                            cell.value = format_period_label(value)
                    else:
                        if is_most_recent_row:
                            cell.font = italic_font
                        
                        if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
                            cell.value = round(float(value), 4)
                            cell.number_format = "0.0000"

                    if is_country_code_row:
                        cell.fill = summary_fill
                        cell.font = Font(bold=True, italic=True)

            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 20
            for col_num in range(3, len(export_df.columns) + 1):
                ws.column_dimensions[get_column_letter(col_num)].width = 15
            ws.row_dimensions[3].height = 40

    output_path.write_bytes(buffer.getvalue())
    return output_path


def fetch_raw_tables(series_list=DEFAULT_SERIES, countries=None, start_period=None, end_period=None):
    client = sdmx.Client("IMF_DATA")
    country_mapping = load_country_mapping()
    tables = {}
    for series in series_list:
        frame = fetch_series_frame(client, series, countries, start_period, end_period)
        tables[series] = pivot_raw_data(frame, country_mapping)
    return tables


def parse_countries(value):
    if not value:
        return None
    return [country.strip().upper() for country in value.split(",") if country.strip()]


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Export raw IMF CPI and PPI index data to Excel.")
    parser.add_argument("-o", "--output", help="Excel output path. Defaults to the latest data period.")
    parser.add_argument("--countries", help="Comma-separated ISO3 country codes. Defaults to all IMF countries.")
    parser.add_argument("--start-period", help='Optional IMF start period, e.g. "2000-M01" or "2000".')
    parser.add_argument("--end-period", help='Optional IMF end period, e.g. "2026-M03" or "2026".')
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    countries = parse_countries(args.countries)
    country_mapping = load_country_mapping()
    tables = fetch_raw_tables(
        countries=countries,
        start_period=args.start_period,
        end_period=args.end_period,
    )

    # Fetch and integrate BLS PPI data
    try:
        bls_frame = fetch_bls_series(
            BLS_PPI_SERIES,
            start_period=args.start_period,
            end_period=args.end_period,
        )
        tables[BLS_PPI_SERIES] = pivot_raw_data(bls_frame, country_mapping)
    except Exception as e:
        print(f"Warning: Could not fetch BLS data: {e}")

    output_path = write_raw_excel(tables, args.output, country_mapping=country_mapping)
    print(f"Wrote {output_path}")
    return output_path


if __name__ == "__main__":
    main()

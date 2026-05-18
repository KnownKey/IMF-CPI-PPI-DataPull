from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from app import (
    DEFAULT_SERIES,
    add_most_recent_row,
    build_key,
    default_output_path,
    format_period_label,
    parse_countries,
    pivot_raw_data,
    write_raw_excel,
)


def test_build_key_defaults_to_all_countries():
    assert build_key("CPI._T.SRP_IX.M") == ".CPI._T.SRP_IX.M"


def test_build_key_filters_selected_countries():
    assert build_key("PPI.IX.M", ["USA", "CAN"]) == "USA+CAN.PPI.IX.M"


def test_parse_countries_normalizes_comma_separated_codes():
    assert parse_countries("usa, can,GBR") == ["USA", "CAN", "GBR"]


def test_format_period_label_formats_imf_months_and_years():
    assert format_period_label("2026-M03") == "March 2026"
    assert format_period_label("2026") == "2026"


def test_pivot_raw_data_keeps_latest_period_first():
    frame = pd.DataFrame(
        {
            "TIME_PERIOD": ["2025-M12", "2026-M01", "2025-M12", "2026-M01"],
            "COUNTRY": ["USA", "USA", "CAN", "CAN"],
            "value": [144.0, 145.0, 130.0, 131.0],
        }
    )

    result = pivot_raw_data(frame)

    assert list(result.index) == ["2026-M01", "2025-M12"]
    assert list(result.columns) == ["CAN", "USA"]
    assert result.loc["2026-M01", "USA"] == 145.0


def test_pivot_raw_data_uses_iso_codes_as_columns():
    frame = pd.DataFrame(
        {
            "TIME_PERIOD": ["2026-M01", "2026-M01"],
            "COUNTRY": ["USA", "CAN"],
            "value": [145.0, 131.0],
        }
    )

    result = pivot_raw_data(frame, {"USA": "United States", "CAN": "Canada"})

    assert list(result.columns) == ["CAN", "USA"]


def test_add_most_recent_row_uses_latest_non_empty_period_per_country():
    df = pd.DataFrame(
        {
            "CAN": [131.0, 130.0, 129.0],
            "USA": [None, 145.0, 144.0],
        },
        index=["2026-M02", "2026-M01", "2025-M12"],
    )

    result = add_most_recent_row(df)

    assert result.loc["Most Recent Data", "CAN"] == "February 2026"
    assert result.loc["Most Recent Data", "USA"] == "January 2026"


def test_default_output_path_uses_current_date():
    tables = {
        DEFAULT_SERIES[0]: pd.DataFrame({"USA": [1.0]}, index=["2026-M01"]),
    }
    date_stamp = datetime.now().strftime("%Y%m%d")
    assert default_output_path(tables).name == f"IMF_CPI_PPI_Raw_{date_stamp}.xlsx"


def test_write_raw_excel_preserves_expected_workbook_formatting(tmp_path):
    tables = {
        DEFAULT_SERIES[0]: pd.DataFrame(
            {"USA": [145.123456, 144.0], "CAN": [131.0, 130.0]},
            index=["2026-M01", "2025-M12"],
        ),
        DEFAULT_SERIES[1]: pd.DataFrame(
            {"USA": [150.0]},
            index=["2026-M01"],
        ),
    }
    output_path = tmp_path / "raw.xlsx"
    mapping = {"USA": "United States", "CAN": "Canada"}

    write_raw_excel(tables, output_path, generated_at=datetime(2026, 5, 15), country_mapping=mapping)
    workbook = load_workbook(output_path)
    ws = workbook["CPI Raw Data"]

    assert ws["A1"].font.italic is True
    assert "latest observation period in this sheet: January 2026" in ws["A1"].value
    assert ws["A2"].font.italic is True
    assert ws["A3"].value is None
    assert ws["B3"].value is None
    assert ws["C3"].value == "United States"  # Row 3 is Name
    assert ws["D3"].value == "Canada"

    assert ws["A4"].value is None
    assert ws["B4"].value == "Country Code"
    assert ws["C4"].value == "USA"  # Row 4 is Code
    assert ws["D4"].value == "CAN"

    assert ws["A5"].value is None
    assert ws["B5"].value == "Most Recent Data"
    assert ws["C5"].value == "January 2026"

    assert ws["A6"].value == 2026
    assert ws["B6"].value == "January 2026"
    assert ws["C6"].number_format == "0.0000"
    assert ws["C6"].value == 145.1235

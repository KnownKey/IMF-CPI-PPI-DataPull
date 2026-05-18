import json
from datetime import datetime
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook
from app import BLS_PPI_SERIES, bls_year_ranges, fetch_bls_series, pivot_raw_data, write_raw_excel


@patch("requests.post")
def test_fetch_bls_series_success(mock_post):
    # Mock successful response
    mock_response = MagicMock()
    mock_response.status_code = 200
    mock_response.json.return_value = {
        "status": "REQUEST_SUCCEEDED",
        "Results": {
            "series": [
                {
                    "seriesID": "WPU00000000",
                    "data": [
                        {"year": "2024", "period": "M01", "value": "123.4"},
                        {"year": "2023", "period": "M12", "value": "122.5"},
                    ],
                }
            ]
        },
    }
    mock_post.return_value = mock_response

    result = fetch_bls_series(BLS_PPI_SERIES, start_period="2023-M01", end_period="2024-M01")

    assert len(result) == 2
    assert result.iloc[0]["TIME_PERIOD"] == "2024-M01"
    assert result.iloc[0]["COUNTRY"] == "USA"
    assert result.iloc[0]["value"] == 123.4
    assert result.iloc[1]["TIME_PERIOD"] == "2023-M12"
    assert result.iloc[1]["value"] == 122.5

    # Verify payload
    args, kwargs = mock_post.call_args
    payload = json.loads(kwargs["data"])
    assert payload["seriesid"] == ["WPU00000000"]
    assert payload["startyear"] == "2023"
    assert payload["endyear"] == "2024"


def test_bls_year_ranges_cover_all_commodities_history():
    assert bls_year_ranges(end_period="2026-M03") == [
        (1913, 1932),
        (1933, 1952),
        (1953, 1972),
        (1973, 1992),
        (1993, 2012),
        (2013, 2026),
    ]


@patch("requests.post")
def test_fetch_bls_series_requests_all_year_chunks(mock_post):
    def response_for_request(url, data, headers):
        payload = json.loads(data)
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {
            "status": "REQUEST_SUCCEEDED",
            "Results": {
                "series": [
                    {
                        "seriesID": "WPU00000000",
                        "data": [
                            {
                                "year": payload["startyear"],
                                "period": "M01",
                                "value": "100.0",
                            }
                        ],
                    }
                ]
            },
        }
        return mock_response

    mock_post.side_effect = response_for_request

    result = fetch_bls_series(BLS_PPI_SERIES, end_period="1955-M12")

    requested_ranges = [
        (
            json.loads(call.kwargs["data"])["startyear"],
            json.loads(call.kwargs["data"])["endyear"],
        )
        for call in mock_post.call_args_list
    ]
    assert requested_ranges == [("1913", "1932"), ("1933", "1952"), ("1953", "1955")]
    assert result["TIME_PERIOD"].tolist() == ["1913-M01", "1933-M01", "1953-M01"]


@patch("requests.post")
def test_fetch_bls_series_error(mock_post):
    # Mock error response
    mock_response = MagicMock()
    mock_response.status_code = 200
    mock_response.json.return_value = {
        "status": "REQUEST_FAILED",
        "message": ["Daily threshold reached"],
    }
    mock_post.return_value = mock_response

    with pytest.raises(RuntimeError, match=r"BLS API error: \['Daily threshold reached'\]"):
        fetch_bls_series(BLS_PPI_SERIES)


def test_bls_data_pivots_correctly():
    bls_frame = pd.DataFrame(
        [
            {"TIME_PERIOD": "2024-M01", "COUNTRY": "USA", "value": 123.4},
            {"TIME_PERIOD": "2023-M12", "COUNTRY": "USA", "value": 122.5},
        ]
    )

    result = pivot_raw_data(bls_frame, {"USA": "United States"})

    assert "USA" in result.columns
    assert result.loc["2024-M01", "USA"] == 123.4
    assert result.index[0] == "2024-M01"


def test_write_raw_excel_handles_bls_source(tmp_path):
    tables = {
        BLS_PPI_SERIES: pd.DataFrame(
            {"United States (USA)": [123.4]},
            index=["2024-M01"],
        ),
    }
    output_path = tmp_path / "bls_raw.xlsx"

    write_raw_excel(tables, output_path, generated_at=datetime(2026, 5, 17))
    workbook = load_workbook(output_path)
    ws = workbook["BLS PPI Raw Data"]

    assert "Raw BLS PPI index data from BLS" in ws["A1"].value
    assert "https://www.bls.gov/ppi/" in ws["A1"].value
    assert "Values are unrebased BLS observations" in ws["A2"].value
